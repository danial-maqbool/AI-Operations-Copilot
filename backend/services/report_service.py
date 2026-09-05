import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session

from backend.models.all_models import (
    Report, Workspace, DataSourceTable, OperationsException, ActionItem, AuditEvent
)
from backend.schemas.report import ReportGenerateRequest
from backend.services.kpi_service import KPIService
from backend.services.entity_views import EntityViewService

EXPORTS_DIR = os.path.join(os.getcwd(), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

class ReportService:
    @staticmethod
    def get_or_create_default_workspace(db: Session) -> str:
        ws = db.query(Workspace).first()
        if not ws:
            ws = Workspace(name="Default Workspace")
            db.add(ws)
            db.commit()
            db.refresh(ws)
        return ws.id

    @staticmethod
    def generate_report(db: Session, request: ReportGenerateRequest) -> Report:
        workspace_id = ReportService.get_or_create_default_workspace(db)
        timestamp_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        base_name = f"opspilot_report_{timestamp_str}"

        # 1. Collect Data
        tables = db.query(DataSourceTable).all()
        avg_health = round(sum(t.data_health_score or 100.0 for t in tables) / len(tables), 1) if tables else 100.0

        kpi_list = KPIService.get_all_kpi_snapshots(db) if request.include_kpis else []
        
        exceptions = db.query(OperationsException).filter(OperationsException.status == "OPEN").all() if request.include_exceptions else []
        exc_data = [
            {
                "id": e.id,
                "title": e.title,
                "severity": e.severity,
                "entity_type": e.entity_type,
                "financial_impact": e.financial_impact,
                "priority_score": e.priority_score,
                "status": e.status,
                "owner": e.owner
            } for e in exceptions
        ]

        actions = db.query(ActionItem).all() if request.include_actions else []
        action_data = [
            {
                "id": a.id,
                "title": a.title,
                "priority": a.priority,
                "owner": a.owner,
                "status": a.status,
                "action_type": a.action_type,
                "due_date": a.due_date.isoformat() if a.due_date else None
            } for a in actions
        ]

        sla_data = EntityViewService.get_sla_risk_monitor() if request.include_sla else {"summary": {}, "breached": [], "at_risk": []}

        sections = {
            "metadata": {
                "generated_at": datetime.utcnow().isoformat(),
                "period": request.period,
                "report_type": request.report_type,
                "data_health_score": avg_health
            },
            "kpis": kpi_list,
            "exceptions": exc_data,
            "actions": action_data,
            "sla": sla_data
        }

        # 2. Generate Excel (.xlsx) Report with OpenPyXL
        xlsx_path = os.path.join(EXPORTS_DIR, f"{base_name}.xlsx")
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        border_side = Side(style='thin', color="CBD5E1")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Tab 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1["A1"] = "OpsPilot — Operations Executive Report"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} | Period: {request.period} | Type: {request.report_type}"
        ws1["A2"].font = sub_font

        ws1["A4"] = "Operational KPI"
        ws1["B4"] = "Value / Status"
        ws1["A4"].font = header_font
        ws1["A4"].fill = header_fill
        ws1["B4"].font = header_font
        ws1["B4"].fill = header_fill

        summary_rows = [
            ("Overall Data Health Score", f"{avg_health}%"),
            ("Active Open Exceptions", str(len(exc_data))),
            ("Open Action Items", str(len([a for a in action_data if a['status'] != 'COMPLETED']))),
            ("SLA Breached Entities", str(sla_data.get("summary", {}).get("breached_count", 0))),
            ("SLA At-Risk Entities", str(sla_data.get("summary", {}).get("at_risk_count", 0))),
            ("Total SLA Exposure", f"${sla_data.get('summary', {}).get('financial_exposure', 0):,.2f}")
        ]
        for idx, (label, val) in enumerate(summary_rows, start=5):
            ws1[f"A{idx}"] = label
            ws1[f"B{idx}"] = val
            ws1[f"A{idx}"].border = cell_border
            ws1[f"B{idx}"].border = cell_border
        
        ws1.column_dimensions["A"].width = 32
        ws1.column_dimensions["B"].width = 25

        # Tab 2: KPIs
        ws2 = wb.create_sheet(title="Key Performance Indicators")
        kpi_headers = ["Code", "Name", "Current Value", "Previous Value", "Change %", "Target", "Status", "Owner"]
        for col_idx, h in enumerate(kpi_headers, start=1):
            c = ws2.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
        
        for r_idx, k in enumerate(kpi_list, start=2):
            ws2.cell(row=r_idx, column=1, value=k.get("code", "")).border = cell_border
            ws2.cell(row=r_idx, column=2, value=k.get("name", "")).border = cell_border
            ws2.cell(row=r_idx, column=3, value=k.get("current_value", 0.0)).border = cell_border
            ws2.cell(row=r_idx, column=4, value=k.get("previous_value", 0.0)).border = cell_border
            ws2.cell(row=r_idx, column=5, value=f"{k.get('pct_change', 0.0):+.1f}%").border = cell_border
            ws2.cell(row=r_idx, column=6, value=k.get("target_value", 0.0)).border = cell_border
            ws2.cell(row=r_idx, column=7, value=k.get("status", "GOOD")).border = cell_border
            ws2.cell(row=r_idx, column=8, value=k.get("owner", "")).border = cell_border

        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws2.column_dimensions[col].width = 20

        # Tab 3: Exceptions
        ws3 = wb.create_sheet(title="Operational Exceptions")
        exc_headers = ["ID", "Title", "Severity", "Entity Type", "Financial Impact ($)", "Priority Score", "Status", "Owner"]
        for col_idx, h in enumerate(exc_headers, start=1):
            c = ws3.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
        
        for r_idx, e in enumerate(exc_data, start=2):
            ws3.cell(row=r_idx, column=1, value=e["id"][:8]).border = cell_border
            ws3.cell(row=r_idx, column=2, value=e["title"]).border = cell_border
            ws3.cell(row=r_idx, column=3, value=e["severity"]).border = cell_border
            ws3.cell(row=r_idx, column=4, value=e["entity_type"]).border = cell_border
            ws3.cell(row=r_idx, column=5, value=e["financial_impact"]).border = cell_border
            ws3.cell(row=r_idx, column=6, value=e["priority_score"]).border = cell_border
            ws3.cell(row=r_idx, column=7, value=e["status"]).border = cell_border
            ws3.cell(row=r_idx, column=8, value=e["owner"]).border = cell_border

        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws3.column_dimensions[col].width = 22

        # Tab 4: Priority Actions
        ws4 = wb.create_sheet(title="Action Items")
        act_headers = ["Action ID", "Title", "Priority", "Owner", "Status", "Action Type", "Due Date"]
        for col_idx, h in enumerate(act_headers, start=1):
            c = ws4.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill

        for r_idx, a in enumerate(action_data, start=2):
            ws4.cell(row=r_idx, column=1, value=a["id"][:8]).border = cell_border
            ws4.cell(row=r_idx, column=2, value=a["title"]).border = cell_border
            ws4.cell(row=r_idx, column=3, value=a["priority"]).border = cell_border
            ws4.cell(row=r_idx, column=4, value=a["owner"]).border = cell_border
            ws4.cell(row=r_idx, column=5, value=a["status"]).border = cell_border
            ws4.cell(row=r_idx, column=6, value=a["action_type"]).border = cell_border
            ws4.cell(row=r_idx, column=7, value=str(a["due_date"] or "")).border = cell_border

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws4.column_dimensions[col].width = 22

        wb.save(xlsx_path)

        # 3. Generate JSON Export
        json_path = os.path.join(EXPORTS_DIR, f"{base_name}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(sections, f, indent=2)

        # 4. Generate CSV Export (Exceptions + KPIs)
        csv_path = os.path.join(EXPORTS_DIR, f"{base_name}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["REPORT_TYPE", "EXECUTIVE_OPERATIONS_REPORT"])
            writer.writerow(["GENERATED_AT", sections["metadata"]["generated_at"]])
            writer.writerow(["DATA_HEALTH_SCORE", f"{avg_health}%"])
            writer.writerow([])
            writer.writerow(["--- KEY PERFORMANCE INDICATORS ---"])
            writer.writerow(["Code", "Name", "Current Value", "Previous Value", "Status", "Owner"])
            for k in kpi_list:
                writer.writerow([k.get("code"), k.get("name"), k.get("current_value"), k.get("previous_value"), k.get("status"), k.get("owner")])
            writer.writerow([])
            writer.writerow(["--- OPERATIONAL EXCEPTIONS ---"])
            writer.writerow(["ID", "Title", "Severity", "Entity", "Financial Impact", "Priority Score", "Status", "Owner"])
            for e in exc_data:
                writer.writerow([e["id"][:8], e["title"], e["severity"], e["entity_type"], e["financial_impact"], e["priority_score"], e["status"], e["owner"]])

        # 5. Save Report Record to Database
        report = Report(
            workspace_id=workspace_id,
            title=request.title,
            period=request.period,
            report_type=request.report_type,
            sections=sections,
            export_formats=["xlsx", "json", "csv"],
            file_path=xlsx_path
        )
        db.add(report)

        # Audit Event
        audit = AuditEvent(
            workspace_id=workspace_id,
            event_type="report_generated",
            entity_type="report",
            entity_id=report.id,
            details={
                "title": request.title,
                "period": request.period,
                "xlsx_file": os.path.basename(xlsx_path)
            }
        )
        db.add(audit)
        db.commit()
        db.refresh(report)

        return report

    @staticmethod
    def list_reports(db: Session, limit: int = 50) -> List[Report]:
        return db.query(Report).order_by(Report.created_at.desc()).limit(limit).all()

    @staticmethod
    def get_report(db: Session, report_id: str) -> Optional[Report]:
        return db.query(Report).filter(Report.id == report_id).first()
